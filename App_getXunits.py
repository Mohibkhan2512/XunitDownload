# coding=utf-8
import sys
sys.coinit_flags = 2  # COINIT_APARTMENTTHREADED

import pyautogui
import time
from tkinter import *
import tkinter as tk
from tkinter import messagebox, ttk
import customtkinter
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import ttkbootstrap as ttk
from PIL import Image
import os
import sys
import shutil
import openpyxl
from tkinter.filedialog import askdirectory, asksaveasfilename
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pathlib import Path


# python -m pip install pywinauto --trusted-host=pypi.python.org --trusted-host=pypi.org --trusted-host=files.pythonhosted.org

# Selecting GUI theme - dark, light , system (for system default) 
customtkinter.set_appearance_mode("light") 
  
# Selecting color theme - blue, green, dark-blue 
customtkinter.set_default_color_theme("green")

global frames
frames = list()
# global activeProcess
activeProcess = 'Null'

class splashScreen:
    def __init__(self, parent, message, waitTime):
        # tk.Toplevel.__init__(self, parent)
        # self.update_idletasks()
        # width = self.winfo_width()
        # frm_width = self.winfo_rootx() - self.winfo_x()
        # win_width = width + 2 * frm_width
        # height = self.winfo_height()
        # titlebar_height = self.winfo_rooty() - self.winfo_y()
        # win_height = height + titlebar_height + frm_width
        # x = self.winfo_screenwidth() // 2 - win_width // 2
        # y = self.winfo_screenheight() // 2 - win_height // 2
        # self.geometry('{}x{}+{}+{}'.format(400, 100, x, y))
        # self.deiconify()

        # self.resizable(0,0)
        # self.configure(bg="white", border=0.5, relief=SOLID)
        # self.wm_overrideredirect(1)
        # self.grab_set()

        self.currentFiltersWindow = tk.Toplevel()
        self.currentFiltersWindow.attributes('-alpha', 0)
        self.currentFilterWindowOpened = self.currentFiltersWindow
        self.currentFiltersWindow.resizable(0,0) 
        self.currentFiltersWindow.configure(bg="white", border=0.5, relief=SOLID)
        self.currentFiltersWindow.wm_overrideredirect(1)
        self.currentFiltersWindow.grab_set()
        # self.currentFiltersWindow.wm_title('Summary of tests')

        def resource_path(relative_path):
            """ Get absolute path to resource, works for dev and for PyInstaller """
            try:
                # PyInstaller creates a temp folder and stores path in _MEIPASS
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, relative_path)

        # self.astriskImg = ImageTk.PhotoImage(Image.open(resource_path("astrisk.png")))

        # self.currentFiltersWindow.iconbitmap(resource_path('icon.ico'))
        # self.currentFiltersWindow.configure(bg="white")
        # self.currentFiltersWindow.geometry("350x340+400+200")

        self.currentFiltersWindow.update_idletasks()
        width = self.currentFiltersWindow.winfo_width()
        frm_width = self.currentFiltersWindow.winfo_rootx() - self.currentFiltersWindow.winfo_x()
        win_width = width + 2 * frm_width
        height = self.currentFiltersWindow.winfo_height()
        titlebar_height = self.currentFiltersWindow.winfo_rooty() - self.currentFiltersWindow.winfo_y()
        win_height = height + titlebar_height + frm_width
        x = self.currentFiltersWindow.winfo_screenwidth() // 2 - win_width // 2
        y = self.currentFiltersWindow.winfo_screenheight() // 2 - win_height // 2
        x -= 100
        self.currentFiltersWindow.geometry('{}x{}+{}+{}'.format(400, 100, x, y))
        self.currentFiltersWindow.attributes('-alpha', 1)
        self.currentFiltersWindow.deiconify()

        # self.geometry("400x100+750+400")

        frame = customtkinter.CTkFrame(self.currentFiltersWindow, border_width=0, corner_radius=0, fg_color="white", bg_color="black")
        frame.place(x=0,y=0, relwidth=1, relheight=1)

        logoContainerFrame = customtkinter.CTkFrame(frame, border_width=0, corner_radius=0, fg_color="white", bg_color="black")#0d6efd
        # logoContainerFrame.place(relx=0.15, rely=0.4)
        logoContainerFrame.place(relwidth=1, relx=0, rely=0.4)


        l1=customtkinter.CTkLabel(logoContainerFrame,text=message, justify="center", text_color="black", bg_color="white")
        # lst1=('Calibri (Body)',18,'bold')
        lst1=('Simplifica', 20)
        l1.configure(font=lst1)
        # l1.place(x=130,y=88)
        l1.pack(side=LEFT, fill=BOTH, expand=TRUE)

        # self.update()
        # parent.update()

        self.bar(frame, waitTime)

    def bar(self, frame, waitTime):
        loadingContainerFrame = customtkinter.CTkFrame(frame, border_width=0, height=50, corner_radius=0, fg_color="white", bg_color="black")#0d6efd
        loadingContainerFrame.pack(side=BOTTOM, fill=X)

        Style = ttk.Style()
        Style.configure('custom.Horizontal.TProgressbar', background='purple', troughcolor='#0d6efd', troughrelief='flat',
                        borderwidth=0, pbarrelief='flat')
        progress=ttk.Progressbar(loadingContainerFrame, style='primary.Striped.Horizontal.TProgressbar',orient=HORIZONTAL,length=410,mode='determinate',)
        progress.pack(side=TOP, fill=BOTH, expand=TRUE)
        
        self.currentFiltersWindow.update()

        # r=0
        maxWaitTime = waitTime
        for i in range(maxWaitTime):
            progress['value']=i*(100/maxWaitTime)            # if r%10 == 0:

            self.currentFiltersWindow.update()
            self.currentFiltersWindow.after(1000, None)
            # r=r+1
    
        return self.currentFiltersWindow.destroy()

class PlaceholderEntry:
    def __init__(self,master,placeholder='', width=98, placeholdercolor='grey',color='black',**kwargs):
        # self.e = ttk.Entry(master,foreground=placeholdercolor,**kwargs, font=("Roboto", 12), justify='left', width=118)
        self.e = ttk.Entry(master,foreground=placeholdercolor,**kwargs, font=("Roboto", 12), justify='left', width=width)
        self.e.bind('<FocusIn>',self.focus_in)
        self.e.bind('<FocusOut>',self.focus_out)
        # self.e.bind("<Button-3>", lambda event, currentEntryInFocus=self.e : self.manipulateTextContent(event, currentEntryInFocus))
        self.e.insert(0, placeholder)
        self.placeholder = placeholder
        self.placeholdercolor=placeholdercolor
        self.color = color

    def pack(self,side=None,**kwargs):
        self.e.pack(side=side,**kwargs)

    def place(self,side=None,**kwargs):
        self.e.place(side=side,**kwargs)

    def grid(self, row=None, column=None,**kwargs):
        self.e.grid(row=row, column=column,**kwargs)

    def focus_in(self,e):
        if self.e.get() == self.placeholder:
            self.e.delete(0,END)
        self.e.configure(foreground=self.color)

    def focus_out(self,e):
        if self.e.get() == '':
            self.e.configure(foreground=self.placeholdercolor)
            self.e.delete(0,END)
            self.e.insert(0,self.placeholder)

class scrollFrameXY():
    def __init__(self, root, **options):
        Style = ttk.Style()
        # Style.configure('custom.outer.TFrame', background="blue")
        Style.configure('custom.TFrame', background="white")
        outerFrame = ttk.Frame(root)
        # outerFrame.place(relheight=1, relwidth=0.90)
        # Style.configure('custom.TCanvas', background="green")
        canvas = ttk.Canvas(outerFrame, highlightthickness=0, bg="red")
        vsb = ttk.Scrollbar(outerFrame, orient="vertical", command=canvas.yview)
        hsb = ttk.Scrollbar(outerFrame, orient="horizontal", command=canvas.xview)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        canvas.place(relheight=1, relwidth=1)
        
        frame = ttk.Frame(canvas, style="custom.TFrame", width=canvas.winfo_width(), height=canvas.winfo_height())
        frame.pack(fill='both', expand=1)
        # frame.place(relheight=1, relwidth=1)
        wrapFrameId = canvas.create_window((0,0), window=frame, anchor="nw", width=400)
        # canvas.config(yscrollcommand=vsb.set)
        # canvas.config(xscrollcommand=hsb.set)
        canvas.bind("<Configure>", lambda event: self.onFrameConfigure())
        canvas.bind("<Enter>", lambda event: canvas.bind_all("<MouseWheel>", self.on_mouse_wheel)) # on mouse enter
        canvas.bind("<Leave>", lambda event: canvas.unbind_all("<MouseWheel>")) # on mouse leave
        self.outerFrame, self.canvas, self.vsb, self.hsb, self.frame, self.wrapFrameId = outerFrame, canvas, vsb, hsb, frame, wrapFrameId 
    def onFrameConfigure(self):
        canvas = self.canvas
        '''Reset the scroll region to encompass the inner frame'''
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfigure(self.wrapFrameId, width=canvas.winfo_width(), height=canvas.winfo_height())
    def on_mouse_wheel(self, event, scale=3):
        canvas = self.canvas

class ToolTipBase:

    def __init__(self, button):
        self.button = button
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0
        self._id1 = self.button.bind("<Enter>", self.enter)
        self._id2 = self.button.bind("<Leave>", self.leave)
        self._id3 = self.button.bind("<ButtonPress>", self.leave)

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hidetip()

    def schedule(self):
        self.unschedule()
        self.id = self.button.after(500, self.showtip)

    def unschedule(self):
        id = self.id
        self.id = None
        if id:
            self.button.after_cancel(id)

    def showtip(self):
        if self.tipwindow:
            return
        x = self.button.winfo_rootx() + 20
        y = self.button.winfo_rooty() + self.button.winfo_height() + 1
        self.tipwindow = tw = tk.Toplevel(self.button)
        self.tipwindow.configure(bg="blue", border=1, relief=SOLID)
        self.tipwindow.wm_overrideredirect(1)
        self.tipwindow.wm_geometry("+%d+%d" % (x, y))
        self.showcontents()

    def showcontents(self, text="Your text here"):
        # Override this in derived class
        # pass ## print("current text => ", text)
        frame = tk.Frame(self.tipwindow, width=300, height=50)
        frame.pack()
        label = tk.Label(frame, text=text, justify=LEFT, wraplength=300,
                      bg="red", border=1)
        label.pack(padx=5, pady=5, fill=X, expand=TRUE)
        # label.pack()

    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()


class ToolTip(ToolTipBase):

    def __init__(self, button, text):
        ToolTipBase.__init__(self, button)
        self.text = text

    def showcontents(self):
        ToolTipBase.showcontents(self, self.text)


class ListboxToolTip(ToolTipBase):

    def __init__(self, button, items):
        ToolTipBase.__init__(self, button)
        self.items = items

    def showcontents(self):
        listbox = Listbox(self.tipwindow, background="white")
        listbox.pack()
        for item in self.items:
            listbox.insert(END, item)


""""""""""""""""""""""""""""""""""""""
""""""""""""""""""""""""""""""""""""""
""" homeWindow class starts here """
""""""""""""""""""""""""""""""""""""""
""""""""""""""""""""""""""""""""""""""
class homeWindow(customtkinter.CTk):
    def __init__(self, root):
        super().__init__()

        self.style = ttk.Style()
        self.root = root

        self.width, self.height = root.winfo_width(), root.winfo_height()

        self.color = "#0d6efd"
        self.font = customtkinter.CTkFont('sans-serif', 12)

        self.containerFrame = customtkinter.CTkFrame(root, corner_radius=0, height=800, fg_color="#eaeaea")
        global frames
        frames.append(self.containerFrame)

        # Tree view styles
        style = ttk.Style()
        style.configure("Treeview",
        background="white",
        foreground="black",
        rowheight=50, # 50
        fieldbackground="white",
        font=('Roberto', 10),
        padding=100)

        style.layout("Treeview", [
            ('Treeview.treearea', {'sticky': 'nswe'})
        ])

        # Change color of headers
        style.configure("Treeview.Heading", 
            background="white", # "#F3A148"self.menuColor
            foreground="black",
            font=('Roberto', 11, 'bold'),
            relief=GROOVE)
        
        # style.map('Treeview.Heading', background=[('active', "#C7C7C7")])

        style.map('Treeview', background=[('selected', "#663399")], foreground=[('selected', "white")])#C7C7C7

        self.containerFrame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(40, 0))
        self.containerFrame.grid_columnconfigure(0, weight=1)
        self.containerFrame.grid_rowconfigure(0, weight=1)
   

        if self.width > 2200:
            pass # pass ## print(self.width)
            self.mainFrameModify = customtkinter.CTkFrame(self.containerFrame, width=1100, fg_color="white")
            self.mainFrameModify.pack(anchor="n", side="top", fill=BOTH, expand=TRUE, padx=10, pady=10)
        else:
            self.mainFrameModify = customtkinter.CTkFrame(self.containerFrame, width=1100, fg_color="white")
            self.mainFrameModify.pack(anchor="n", side="top", fill=BOTH, expand=TRUE, padx=10, pady=10)
        
        self.mainFrameModifyForSeperator = customtkinter.CTkFrame(self.mainFrameModify, width=1100, height=100, fg_color="white")
        self.mainFrameModifyForSeperator.pack(anchor='n', side="top", fill=BOTH, padx=50, pady=10)
       
        self.mainFrameModifyForTargetData = customtkinter.CTkFrame(self.mainFrameModify, width=1100, fg_color="white")
        self.mainFrameModifyForTargetData.pack(anchor='n', side="top", fill=BOTH, padx=50, pady=50)
        

        self.downloadLable = customtkinter.CTkLabel(self.mainFrameModifyForSeperator, anchor="w", text="Download Template", font=('Arial',25,'bold'))
        self.downloadLable.grid(row=0, column=0, pady=(30, 30))
        # self.downloadLable.place(x=55, y=40, relwidth=1)
# 
        separator = ttk.Separator(self.mainFrameModifyForSeperator, orient='horizontal', style='primary.Horizontal.TSeparator')
        # separator.grid(row=1, column=0, sticky="nsew", columnspan=2)
        # separator.pack(fill='x')
        separator.place(x=0, rely=0.9, relwidth=1)
# 

        # take the data
        lst = [('Template', 'Version', 'Last updated on', ''),
            ('AxiomLogUploadTemplate', 'v1.0', '15/07/2024', 'Click to Download'),
            ('ScriptExecutionTemplate', 'v1.0', '15/07/2024', 'Click to Download')]
        
        # find total number of rows and
        # columns in list
        total_rows = len(lst)
        total_columns = len(lst[0])
        self.allEntries = []

        # code for creating table
        for i in range(total_rows):
            for j in range(total_columns):
                if j == 0:
                    # self.e = ttk.Entry(self.mainFrameModifyForTargetData, width=40, foreground='black',
                    #            font=('Arial',16,'bold'), state="normal")
                    self.e = customtkinter.CTkTextbox(self.mainFrameModifyForTargetData, corner_radius=0, fg_color="grey", text_color="black",border_spacing=0, border_color="grey",
                                                            border_width=0,  width=600, wrap=WORD, height=120, bg_color="blue", font=('Arial',20,'bold'))
                    # .CTkTextbox()
                        
                else:
                    # self.e = ttk.Entry(self.mainFrameModifyForTargetData, width=20, foreground='black',
                    #            font=('Arial',16,'bold'), state="normal")
                    self.e = customtkinter.CTkTextbox(self.mainFrameModifyForTargetData, corner_radius=0, fg_color="grey", text_color="black",border_spacing=0, border_color="grey",
                                                            border_width=0,  width=400, wrap=WORD, height=120, bg_color="blue", font=('Arial',20,'bold'))
                 
                self.allEntries.append(self.e)
                self.e.tag_config("center", justify="center")

                self.e.grid(row=i+2, column=j)
                self.e.insert(END, "\n\n"+lst[i][j], "center")
        
        self.downloadAxiomTemplate = self.allEntries[7]
        self.downloadAxiomTemplate.bind("<1>", lambda event, templateType='axiom' : self.downloadTemplate(event, templateType))

        self.downloadScriptExecTemplate = self.allEntries[-1]
        self.downloadScriptExecTemplate.bind("<1>", lambda event, templateType='script' : self.downloadTemplate(event, templateType))


        for index, entry in enumerate(self.allEntries):
            # print(f"index => {index} and content => {entry.get('0.0', END)}")
            if index  < 4:
                entry.configure(fg_color="#0d6efd", border_color="grey", text_color="white", state="disabled", font=('Arial',20,'bold'))
            else:
                entry.configure(fg_color="white", border_color="white", text_color="black", state="disabled")
        
        self.allEntries[-1].configure(fg_color="#0d6efd", border_color="grey", text_color="white", state="disabled", font=('Arial',20,'bold'))
        self.allEntries[7].configure(fg_color="#0d6efd", border_color="grey", text_color="white", state="disabled", font=('Arial',20,'bold'))

    def downloadTemplate(self, event, templateType):
        print("downloadTemplate called")
        print("templateType => ", templateType)
        if templateType == "axiom":
            dlg = asksaveasfilename(confirmoverwrite=False)
            fname = dlg
            if fname != '':
                shutil.copyfile(resource_path('Templates\\AxiomLogUploadTemplate.xlsx'), fname+'.xlsx')  
        else:
            dlg = asksaveasfilename(confirmoverwrite=False)
            fname = dlg
            if fname != '':
                shutil.copyfile(resource_path('Templates\\ScriptExecutionTemplate.xlsx'), fname+'.xlsx')  


""""""""""""""""""""""""""""""""""""""
""""""""""""""""""""""""""""""""""""""
""" Upload logs to axiom class starts here """
""""""""""""""""""""""""""""""""""""""
""""""""""""""""""""""""""""""""""""""
class getXunitFiles(customtkinter.CTk):
    def __init__(self, root):
        super().__init__()

        # global declarations
        self.downloadDoneFor = []
        self.couldNotDownload = []
        self.skippedLogUpload = []
        self.downloadFailedPair = []
        self.allRecords = []
        self.processedRecords = {}
        self.summaryForJob = {}
        self.currentFilterWindowOpened = None

        self.style = ttk.Style()
        self.root = root

        self.width, self.height = root.winfo_width(), root.winfo_height()

        self.color = "#0963e8"
        self.font = customtkinter.CTkFont('sans-serif', 12)

        self.containerFrame = customtkinter.CTkFrame(root, corner_radius=0, height=800, fg_color="#eaeaea")
        global frames
        frames.append(self.containerFrame)

        # Tree view styles
        style = ttk.Style()
        style.configure("Treeview",
        background="white",
        foreground="black",
        rowheight=50, # 50
        fieldbackground="white",
        font=('Roberto', 10),
        padding=100)

        style.layout("Treeview", [
            ('Treeview.treearea', {'sticky': 'nswe'})
        ])

        # Change color of headers
        style.configure("Treeview.Heading", 
            background="white", # "#F3A148"self.menuColor
            foreground="black",
            font=('Roberto', 11, 'bold'),
            relief=GROOVE)
        
        # style.map('Treeview.Heading', background=[('active', "#C7C7C7")])

        headinStyle = "Treeview.Heading"

        style.map('Treeview', background=[('selected', "#0963e8")], foreground=[('selected', "white")])#C7C7C7

        self.containerFrame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(40, 0))
        self.containerFrame.grid_columnconfigure(0, weight=1)
        self.containerFrame.grid_rowconfigure(0, weight=1)
   

        if self.width > 2200:
            pass # pass ## print(self.width)
            self.mainFrameModify = customtkinter.CTkFrame(self.containerFrame, width=1100, height=1100, fg_color="#eaeaea")
            self.mainFrameModify.pack(anchor="n", side="top", fill=BOTH, expand=TRUE, padx=10, pady=(0, 10))
        else:
            self.mainFrameModify = customtkinter.CTkFrame(self.containerFrame, width=1100, height=800, fg_color="#eaeaea")
            self.mainFrameModify.pack(anchor="n", side="top", fill=BOTH, expand=TRUE, padx=10, pady=(0, 10))
        
       
        self.mainFrameModifyForTargetData = customtkinter.CTkFrame(self.mainFrameModify, width=1100, height=50, fg_color="white")
        self.mainFrameModifyForTargetData.pack(anchor='n', side="top", fill=X, pady=10)

        self.covLinkEntry = PlaceholderEntry(self.mainFrameModifyForTargetData, placeholder="Enter coverage link...")
        self.covLinkEntry = self.covLinkEntry.e

        self.covLinkEntry.configure(style='info.TEntry')

        self.covLinkEntry.pack(anchor='n', side="top", fill=BOTH, pady=(20, 5), padx=(20, 20))

        self.pathToStoreEntry = PlaceholderEntry(self.mainFrameModifyForTargetData, placeholder="Click to select directory path to store downloaded X-unit files...")
        self.pathToStoreEntry = self.pathToStoreEntry.e
        self.storeTestPlanLoc = StringVar()
        self.pathToStoreEntry.configure(style='info.TEntry')
        self.pathToStoreEntry.bind("<ButtonRelease-1>", self.storeTestPlanPath)


        self.pathToStoreEntry.pack(anchor='n', side="top", fill=BOTH, pady=5, padx=(20, 20))

        # Create a frame to hold the entries
        self.entryFrame = customtkinter.CTkFrame(self.mainFrameModifyForTargetData, fg_color="white")
        self.entryFrame.pack(anchor='w', pady=(5, 20))

        # Set the width of the frame to 70% of the parent width
        self.entryFrame.update_idletasks()  # Update the frame to get its current width
        frame_width = int(self.mainFrameModifyForTargetData.winfo_width() * 0.6)
        self.entryFrame.configure(width=frame_width)

        # Create username entry
        self.usernameEntry = PlaceholderEntry(self.entryFrame, placeholder="Enter Username...", width=80)
        self.usernameEntry = self.usernameEntry.e
        self.usernameEntry.configure(style='info.TEntry')
        self.usernameEntry.pack(side="left", expand=True, fill=customtkinter.BOTH, padx=(20, 10))  # Expand to fill available space

        # Create password entry
        self.passwordEntry = PlaceholderEntry(self.entryFrame, placeholder="Enter Password...", width=80)
        self.passwordEntry = self.passwordEntry.e
        self.passwordEntry.configure(style='info.TEntry')
        self.passwordEntry.bind("<FocusIn>", self.configureForPassword)
        self.passwordEntry.bind("<FocusOut>", self.configureAfterPassword)
        self.passwordEntry.pack(side="left", expand=True, fill=customtkinter.BOTH, padx=(10, 20))  # Expand to fill available space

        # Create Start Download button
        self.startLogUploadButton = customtkinter.CTkButton(
            master=self.entryFrame,
            width=170,
            text="Fetch Job IDs",
            corner_radius=4,
            font=customtkinter.CTkFont('Helvetica', 14, "bold"),
            border_spacing=4,
            bg_color="white",
            fg_color="#0d6efd",
            hover_color="#0963e8",
            command=lambda: self.validateInputEntries()
        )
        self.startLogUploadButton.pack(anchor='n', side="left", fill=customtkinter.Y, padx=(0, 20))

        # Create Stop Download button
        self.abortJobButton = customtkinter.CTkButton(
            master=self.entryFrame,
            width=170,
            font=customtkinter.CTkFont('Helvetica', 14, "bold"),
            text="Start XUnit download",
            corner_radius=4,
            border_spacing=4,
            bg_color="white",
            fg_color="#0d6efd",
            hover_color="#0963e8",
            command=lambda: self.startXunitDownload(),
            # state="disabled"
        )
        self.abortJobButton.pack(anchor='n', side="left", fill=customtkinter.Y, padx=(0, 20))

        self.scrollableFrame = scrollFrameXY(self.mainFrameModify)

        self.scrollableFrame.outerFrame.pack(side=TOP, fill="both", expand=1)

        self.treeView = ttk.Treeview(self.scrollableFrame.frame, selectmode='extended', 
                                    yscrollcommand=self.scrollableFrame.vsb.set, 
                                    xscrollcommand=self.scrollableFrame.hsb.set,
                                    height=10)
        
        self.treeView.bind("<Double-1>", self.handleDoubleClick)
        self.treeView.bind("<Button-3>", self.manipulateColumnOfSheet)
        self.treeView.bind("<ButtonRelease-1>", self.disableSelectedRecords)
        self.treeView.bind('<Control-a>', self.selectAll)
        self.treeView.bind('<<TreeviewSelect>>', lambda event: self.itemSelected(event))
        

        self.scrollableFrame.vsb.config(command=self.treeView.yview)
        self.scrollableFrame.hsb.config(command=self.treeView.xview)
        
        self.treeView.pack(pady=(0, 0), side=TOP, fill="both", expand=True)

        self.columns = ['','SrNo.', 'Job Id', 'X-Unit Files Download Summary', '']
        self.columnslowerCase = ['','srno.', 'job id', 'x-unit files download summary', '']

        self.treeView["column"] = self.columns
        self.treeView["show"] = "headings"

        for column in self.treeView["columns"]:
            self.treeView.heading(column, text=column)
        
        for column in self.treeView["columns"]:
            self.treeView.column(column, minwidth=300, width=300, anchor=CENTER, stretch=False)
        
        self.treeView.heading("#0", text="\n")
        self.treeView.column("#1", width=5, minwidth=5, anchor=CENTER, stretch=False)
        self.treeView.column("#2", width=100, minwidth=100, anchor=CENTER, stretch=False)
        self.treeView.column("#3", width=300, minwidth=300, anchor=CENTER, stretch=False)
        self.treeView.column("#4", width=1480, minwidth=1480, anchor=CENTER, stretch=False)
        self.treeView.column("#5", width=5, minwidth=5, anchor=CENTER, stretch=False)
        # self.treeView.column("#4", width=850, minwidth=850, anchor=CENTER, stretch=False)
        # self.treeView.column("#5", width=310, minwidth=310, anchor=CENTER, stretch=False)

    def validateInputEntries(self):
        if (self.covLinkEntry.get() == "" or self.pathToStoreEntry.get() == "" or \
            self.usernameEntry.get() == "" or self.passwordEntry.get() == ""):
            messagebox.showerror("Error", "Please fill all fields to continue...")
            return
        
        # Extract table rows from tbody
        self.jobIDList = []
        self.jobNameList = []

        self.extractJobLinksFromCoverage()
        self.root.after(1000)
        self.root.update()
        self.getXunitsFromCov()


    def extractJobLinksFromCoverage(self):
        self.covLink = self.covLinkEntry.get()
        # storeXunitsAt = self.pathToStoreEntry.get()

        # Set up the WebDriver
        driver = webdriver.Edge()

        # url = 'https://axiom.qualcomm.com/#/jobs/history/search;vs=eyJmIjpbeyJ2YWx1ZSI6IlNBODY1MFAuUVguNC40LjAgRVMxNCIsImtleSI6ImNvdmVyYWdlUGxhbiIsIm1hdGNoTW9kZSI6ImN0In1dLCJpIjoxLCJwcyI6NTB9'
        url = self.covLink.strip()

        print(f'url => {url}')

        username = self.usernameEntry.get().strip()
        password = self.passwordEntry.get().strip()

        try:
            # Open the web page
            driver.get(url)
        except:
            pass

        # Increase the wait time
        wait = WebDriverWait(driver, 10)
        import pyautogui
        import time

        time.sleep(1)

        pyautogui.write(username)
        pyautogui.press('tab')
        pyautogui.write(password)
        pyautogui.press('tab')

        pyautogui.press('enter')
        
        t = 0
        while(t <= 10):
            if (t < 10):
                try:
                    # Wait for the page to load completely
                    wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
                    break
                except:
                    t += 1
                    time.sleep(1)
            else:
                messagebox.showerror("Error", "Invalid username or password, Please try again with valid credentials...")
                return

        # Execute JavaScript to ensure all tables are loaded
        driver.execute_script("return document.readyState == 'complete'")

        # jobIDList = []
        curIdx = 0
        while True:
            try:
                print("In try")
                curIdx+=1

                lastPage = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'p-ripple.p-element.p-paginator-last.p-paginator-element.p-link.ng-star-inserted.p-disabled')))

                tables = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'table')))
                print(f'tables -> {tables}')
                table = tables[0]

                # Extract table headers from thead
                headers = []
                thead = table.find_element(By.TAG_NAME, 'thead')
                for row in thead.find_elements(By.TAG_NAME, 'tr'):
                    for header in row.find_elements(By.TAG_NAME, 'th'):
                        headers.append(header.text.strip())

                print(f'headers => {headers}')

                time.sleep(2)

                rows = []
                tbody = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'tbody')))
                tbody = tbody[0]
                for row in tbody.find_elements(By.TAG_NAME, 'tr'):
                    job = row.find_element(By.TAG_NAME, 'a')
                    self.jobNameList.append(job.text)


                print(f'len of self.jobNameList in try => {len(self.jobNameList)}')

                break
            except:
                print("In except")
                curIdx+=1

                tables = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'table')))
                # print(f'tables -> {tables}')
                table = tables[0]

                # Extract table headers from thead
                headers = []
                thead = table.find_element(By.TAG_NAME, 'thead')
                for row in thead.find_elements(By.TAG_NAME, 'tr'):
                    for header in row.find_elements(By.TAG_NAME, 'th'):
                        headers.append(header.text.strip())

                print(f'headers => {headers}')

                time.sleep(2)

                rows = []
                tbody = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'tbody')))
                tbody = tbody[0]
                for row in tbody.find_elements(By.TAG_NAME, 'tr'):
                    job = row.find_element(By.TAG_NAME, 'a')
                    self.jobNameList.append(job.text)

                print(f'len of self.jobNameList => {len(self.jobNameList)}')
                try:
                    nextPage = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'p-ripple.p-element.p-paginator-next.p-paginator-element.p-link')))
                    nextPage[0].click()
                except:
                    break

                # for job in jobIDList:
                #     print(f'job => {job.text}')

        # Close the WebDriver
        driver.quit()

    def getXunitsFromCov(self):
        print(f'len of jobIDList in getXunitsFromCov=> {len(self.jobNameList)}')

        for job in self.jobNameList:
            print(f'job => {job}')
        
        df = pd.DataFrame(self.jobNameList, columns=['Numbers'])

        indexColumn = []
        blankOne = []
        blankTwo = []
        # df = df.drop([df.index[0], df.index[1]])
        
        for i in range(1, len(df) + 1):
            indexColumn.append(i)
            blankOne.append('')
            blankTwo.append('')
        
        df.insert(0, "blankOne", blankOne, True)
        df.insert(1, "SrNo.", indexColumn, True)
        df.insert(3, "blankTwo", blankTwo, True)
        # print("df => ", df)
        # df.drop(df.index[1], inplace=True)
        dataFrame = df
        dataFrame = dataFrame.fillna('')
        df_rows = dataFrame.to_numpy().tolist()
        print(df_rows)

        self.treeView.tag_configure("oddrow", background='white',foreground='black')
        self.treeView.tag_configure("evenrow", background='#dbeafe',foreground='black') #dbeafe

        for index, row in enumerate(df_rows):
            self.root.update()
            if row[2] != '':
                if index % 2 == 0:
                    self.treeView.insert("", "end", values=row, tags=('evenrow',))
                if index % 2 != 0:
                    self.treeView.insert("", "end", values=row, tags=('oddrow',))
        
        # self.abortJobButton.configure(state="enabled")

    def startXunitDownload(self):
        self.root.update()
        storeXunitsAt = self.pathToStoreEntry.get().strip()
        self.allRecords = []
        treeViewContent = self.treeView.get_children()

        if len(self.allRecords) == 0:
            for record in treeViewContent:
                self.allRecords.append(self.treeView.item(record)['values'])

        if len(self.allRecords) == 0:
            messagebox.showerror("Error", "Please enter Job IDs to start download...")
            return

        print("allRecords => ", self.allRecords)
       
        username = self.usernameEntry.get().strip()
        password = self.passwordEntry.get().strip()

        if self.pathToStoreEntry == "" or self.pathToStoreEntry == "Click to select directory path to store downloaded X-unit files...":
            messagebox.showerror("Error", "Please enter valid directory path to store X-Unit files...")
            return

        if username == "Enter Username..." or username == "" or password == "Enter Password..." or password == "":
            messagebox.showerror("Error", "Please enter valid username and password to continue...")
            return

        for job in self.allRecords:
            job = job[2]
            if job != '':
                axiomUrl = 'https://axiom.qualcomm.com/#/reports/jobs/' + str(job)
                print(f'axiomUrl => ', axiomUrl)
            else:
                continue
        
            # Set up the WebDriver
            driver = webdriver.Edge()

            url = axiomUrl.replace('reports', 'logs')

            print(f'url => {url}')

            try:
                # Open the web page
                driver.get(url)
            except:
                continue

            # Increase the wait time
            wait = WebDriverWait(driver, 120)

            time.sleep(1)

            pyautogui.write(username)
            pyautogui.press('tab')
            pyautogui.write(password)
            pyautogui.press('tab')

            pyautogui.press('enter')

            # Wait for the page to load completely
            wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

            # Execute JavaScript to ensure all tables are loaded
            driver.execute_script("return document.readyState == 'complete'")

                
            tables = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'table')))
            # print(f'tables -> {tables}')
            table = tables[0]

            # Extract table headers from thead
            headers = []
            thead = table.find_element(By.TAG_NAME, 'thead')
            for row in thead.find_elements(By.TAG_NAME, 'tr'):
                for header in row.find_elements(By.TAG_NAME, 'th'):
                    headers.append(header.text.strip())

            print(f'headers => {headers}')

            # Extract table rows from tbody
            rows = []
            tbody = table.find_element(By.TAG_NAME, 'tbody')
            for row in tbody.find_elements(By.TAG_NAME, 'tr'):
                cells = row.find_elements(By.TAG_NAME, 'td')
                if len(cells) > 0:
                    rows.append([cell.text.strip() for cell in cells])

            # Create a DataFrame from the extracted data
            df = pd.DataFrame(rows, columns=headers)

            # print('df => ', df)

            atagsList = df['unfold_more\nName'].astype(str).tolist()

            print(f'Overall Runs for the job {job} => {atagsList}')

            saveDone = 1
            preprocessingFoundFor = 0
            self.summaryForJob[job] = []
            for index, tag in enumerate(atagsList):
                tables = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'table')))
                print(f'tables -> {tables}')
                table = tables[0]

                tbody = table.find_element(By.TAG_NAME, 'tbody')
                for row in tbody.find_elements(By.TAG_NAME, 'tr'):
                    atags = row.find_elements(By.TAG_NAME, 'a')
                    # print(f'atags => {atags}')

                    # atagsListFinal.append(atag)
                    for atag in atags:
                        if atag.text == tag:
                            curAtag = atag
                            break

                # print(f'index => {index}')
                # print(f'tag => {curAtag.text}')
                
                curAtag.click()
                PostprocessingFound = False
                # time.sleep(2)
                # # break
                atagsListForPostProssessing = []
                tables = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'table')))
                # print(f'tables two -> {tables}')

                # print(f'headers two => {headers}')

            
                table = tables[0]

                tbody = table.find_element(By.TAG_NAME, 'tbody')
                for row in tbody.find_elements(By.TAG_NAME, 'tr'):
                    atag = row.find_elements(By.TAG_NAME, 'a')
                    atagsListForPostProssessing.append(atag)


                for i, atag in enumerate(atagsListForPostProssessing):
                    if atag[0].text == 'Postprocessing':
                        self.root.update()
                        PostprocessingFound = True
                        print(f'Postprocessing found for {tag}...')
                        preprocessingFoundFor += 1

                        atag[0].click()

                        # click on xunit folder

                        atagsListForXunit = []
                        tables = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'table')))

                        table = tables[0]

                        tbody = table.find_element(By.TAG_NAME, 'tbody')
                        for row in tbody.find_elements(By.TAG_NAME, 'tr'):
                            atagXunit = row.find_elements(By.TAG_NAME, 'a')
                            atagsListForXunit.append(atagXunit)

                        print(f'atagsListForXunit => {atagsListForXunit}')

                        for atagXunit in atagsListForXunit:
                            if atagXunit[0].text == 'XUnit_Summary-Genarate':
                                print(f'Postprocessing found in Postprocessing for {tag}...')
                                atagXunit[0].click()
                                break

                        checkBoxDiv = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'p-checkbox.p-component')))
                        checkBoxDiv[0].click()

                        time.sleep(0.5)

                        downloadBtn = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'p-element.standard.p-button.p-component.ng-star-inserted')))

                        downloadBtn[1].click()
                        print('downloadBtn clicked...')

                        time.sleep(1)

                        # Open the downloads page in Edge
                        driver.get('edge://downloads/')

                        # Wait for the downloads page to load
                        time.sleep(2)

                        # Wait for the "Keep" button to be clickable
                        # save = 'save'+str(index+1)
                        save = 'save'+str(saveDone)
                        print(f'save state => {save}')
                        keep_button = wait.until(EC.presence_of_all_elements_located((By.ID, save)))
                        saveDone += 1

                        print(f'keep_button {keep_button}')
                        keep_button[0].click()
                        # keep_button.click()
                        print('keep_button clicked...')
                        # except:
                        #     continue

                        bg = time.time()
                        print(f'sleep begin {bg}')
                        print('sleep for 10 sec')
                        time.sleep(5)
                        eg = time.time()
                        print(f'sleep end {eg-bg}')
                        # Define the download directory and the target directory

                        # Get the user's home directory
                        home_dir = str(Path.home())

                        # Construct the path to the download directory
                        if os.name == 'nt':  # For Windows
                            download_dir = os.path.join(home_dir, 'Downloads')

                        print(f"Download directory: {download_dir}")

                        # download_dir = '/path/to/download/directory'
                        target_dir = storeXunitsAt

                        # Get the list of files in the download directory
                        files = os.listdir(download_dir)

                        # Sort files by modification time in descending order
                        files.sort(key=lambda x: os.path.getmtime(os.path.join(download_dir, x)), reverse=True)

                        # Get the most recently downloaded file
                        recent_file = files[0]

                        print(f'recent_file => {recent_file}')

                        if 'xunit_report' in recent_file:
                            # Construct full file paths
                            src_path = os.path.join(download_dir, recent_file)
                            dst_path = os.path.join(target_dir, recent_file)
                            # Move the file
                            shutil.move(src_path, dst_path)
                            self.downloadDoneFor.append(job)

                            print(f"Moved {recent_file} to {target_dir}")
                        break
                    
                    elif atag[0].text == 'XUnit_Summary-Genarate':
                        self.root.update()
                        PostprocessingFound = True
                        print(f'XUnit_Summary found for {tag}...')
                        preprocessingFoundFor += 1
                        
                        atag[0].click()

                        checkBoxDiv = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'p-checkbox.p-component')))
                        checkBoxDiv[0].click()

                        time.sleep(0.5)

                        downloadBtn = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'p-element.standard.p-button.p-component.ng-star-inserted')))

                        downloadBtn[1].click()
                        print('downloadBtn clicked...')

                        # Open the downloads page in Edge
                        driver.get('edge://downloads/')

                        # Wait for the downloads page to load
                        time.sleep(2)

                        # Wait for the "Keep" button to be clickable
                        save = 'save'+str(saveDone)
                        print(f'save state => {save}')
                        keep_button = wait.until(EC.presence_of_all_elements_located((By.ID, save)))
                        saveDone += 1

                        print(f'keep_button {keep_button}')
                        keep_button[0].click()
                        # keep_button.click()
                        print('keep_button clicked...')

                        bg = time.time()
                        print(f'sleep begin {bg}')
                        print('sleep for 10 sec')
                        time.sleep(5)
                        eg = time.time()
                        print(f'sleep end {eg-bg}')
                        # Define the download directory and the target directory

                        # Get the user's home directory
                        home_dir = str(Path.home())

                        # Construct the path to the download directory
                        if os.name == 'nt':  # For Windows
                            download_dir = os.path.join(home_dir, 'Downloads')

                        print(f"Download directory: {download_dir}")

                        # download_dir = '/path/to/download/directory'
                        target_dir = storeXunitsAt

                        # Get the list of files in the download directory
                        files = os.listdir(download_dir)

                        # Sort files by modification time in descending order
                        files.sort(key=lambda x: os.path.getmtime(os.path.join(download_dir, x)), reverse=True)

                        # Get the most recently downloaded file
                        recent_file = files[0]

                        print(f'recent_file => {recent_file}')

                        if 'xunit_report' in recent_file:
                            # Construct full file paths
                            src_path = os.path.join(download_dir, recent_file)
                            dst_path = os.path.join(target_dir, recent_file)
                            # Move the file
                            shutil.move(src_path, dst_path)
                            self.downloadDoneFor.append(job)

                            print(f"Moved {recent_file} to {target_dir}")
                        break
                
                if PostprocessingFound == False:
                    self.couldNotDownload.append(job)
                    print(f'Postprocessing not found for {tag}...')
                
                # self.summaryForJob[job] = []
                self.summaryForJob[job].append([tag, PostprocessingFound])
                self.root.update()
                self.root.after(1000, None)
                # print(f'url -> {url}')
                driver.get(url)
                # Press Alt + D to select browser address bar
                pyautogui.hotkey('alt', 'd')
                pyautogui.press('enter')

                time.sleep(2)

            self.processedRecords[job] = []
            self.processedRecords[job].append([len(atagsList), preprocessingFoundFor])
            self.refreshTreeView()
            self.root.update()
            self.root.after(1000, None)
            # Close the WebDriver
            driver.quit()

    def refreshTreeView(self):
        print("refreshTreeView called")

        print(f"self.processedRecords => {self.processedRecords} ")

        self.treeView.delete(*self.treeView.get_children())
        self.treeView.tag_configure("uploadDoneRow", background='#D3D3D3',foreground='#606060')
        self.treeView.tag_configure("couldNotUploadEven", background='#ffb09c',foreground='#606060')
        self.treeView.tag_configure("couldNotUploadOdd", background='#fbd9d3',foreground='#606060')
        
        for index, row in enumerate(self.allRecords):
            flag = False
            flagTwo = False 
            xunitMismatch = False

            try:
                data = self.processedRecords[row[2]]
                print(f"data => {data}")
                recordData = f"Found total {data[0][1]} X-Unit files for {data[0][0]} runs of the Job  | |  Double click to see summary"
                row[3] = recordData
                if data[0][1] != data[0][0]:
                    xunitMismatch = True
            except:
                pass

            print(f"Row => {row}")

            if row[2] in self.couldNotDownload or row in self.couldNotDownload and xunitMismatch:
                if (index % 2) == 0:
                    self.treeView.insert("", "end", values=row, tags=('couldNotUploadEven',))
                    flagTwo = True
                else:
                    self.treeView.insert("", "end", values=row, tags=('couldNotUploadOdd',))
                    flagTwo = True
                continue

            if row[2] in self.downloadDoneFor and flagTwo == False:
                self.treeView.insert("", "end", values=row, tags=('uploadDoneRow',))
                flag = True
                continue           

            if index % 2 == 0 and flag == False and flagTwo == False:
                self.treeView.insert("", "end", values=row, tags=('evenrow',))
            if index % 2 != 0 and flag == False and flagTwo == False:
                self.treeView.insert("", "end", values=row, tags=('oddrow',))

    def viewJobSummary(self, jobSummary, jobIdName):
        if self.currentFilterWindowOpened is not None:
            self.currentFilterWindowOpened.destroy()
        
        self.currentFiltersWindow = tk.Toplevel()
        self.currentFilterWindowOpened = self.currentFiltersWindow
        self.currentFiltersWindow.resizable(1,0) 
        self.currentFiltersWindow.wm_title(f'Summary of X-Unit Downloads for job: {jobIdName}')

        def resource_path(relative_path):
            """ Get absolute path to resource, works for dev and for PyInstaller """
            try:
                # PyInstaller creates a temp folder and stores path in _MEIPASS
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, relative_path)

        # self.astriskImg = ImageTk.PhotoImage(Image.open(resource_path("astrisk.png")))

        self.currentFiltersWindow.iconbitmap(resource_path('icon.ico'))
        self.currentFiltersWindow.configure(bg="white")
        # self.currentFiltersWindow.geometry("350x340+400+200")

        self.currentFiltersWindow.update_idletasks()
        width = self.currentFiltersWindow.winfo_width()
        frm_width = self.currentFiltersWindow.winfo_rootx() - self.currentFiltersWindow.winfo_x()
        win_width = width + 2 * frm_width
        height = self.currentFiltersWindow.winfo_height()
        titlebar_height = self.currentFiltersWindow.winfo_rooty() - self.currentFiltersWindow.winfo_y()
        win_height = height + titlebar_height + frm_width
        x = self.currentFiltersWindow.winfo_screenwidth() // 2 - win_width // 2
        y = self.currentFiltersWindow.winfo_screenheight() // 2 - win_height // 2
        x -= 100
        self.currentFiltersWindow.geometry('{}x{}+{}+{}'.format(700, 360, x, y))
        self.currentFiltersWindow.deiconify()

        self.currentFiltersFrame = customtkinter.CTkScrollableFrame(self.currentFiltersWindow, fg_color="white", 
                                                                    scrollbar_button_color="white", scrollbar_button_hover_color="gray",
                                                                    border_width=0.1, border_color="#eaeaea")
        self.currentFiltersFrame.pack(padx=5,pady=5, side=TOP, fill="both", expand=1)



        filtersContainerFrame = customtkinter.CTkFrame(self.currentFiltersFrame, border_width=0.1, corner_radius=1, fg_color=self.color, height=250)
        filtersContainerFrame.pack(side=TOP, fill=BOTH, padx=(10,0), pady=(10,0))

        nameOfColumn = customtkinter.CTkTextbox(master=filtersContainerFrame, corner_radius=0, width=300, fg_color=self.color, text_color="white", border_spacing=0, border_color=self.color,
                                                        border_width=0, height=1, bg_color="blue", font=('Helvetica', 12, 'bold')) # , style="success.TEntry"
                    
        nameOfColumn.insert(END, "Job Runs")
        nameOfColumn.configure(state=DISABLED)
        nameOfColumn.pack(side=LEFT, fill=X)
        
        filterOnColumn = customtkinter.CTkTextbox(master=filtersContainerFrame, corner_radius=0, width=2000, fg_color=self.color, text_color="white", border_spacing=0, border_color=self.color,
                                                        border_width=0, height=1, bg_color="blue", font=('Helvetica', 12, 'bold')) # , style="success.TEntry"
                    
        filterOnColumn.insert(END, "X-Unit found")
        filterOnColumn.configure(state=DISABLED)
        filterOnColumn.pack(side=LEFT, fill=X)

        # filtersContainerFixFrame = customtkinter.CTkFrame(self.currentFiltersFrame, border_width=0.1, corner_radius=4, fg_color="#eaeaea")
        # filtersContainerFixFrame.pack(side=TOP, fill=X, padx=(10,0))

        for job in jobSummary:
            filtersContainerFixFrame = customtkinter.CTkFrame(self.currentFiltersFrame, border_width=0.1, corner_radius=4, fg_color="#eaeaea")
            filtersContainerFixFrame.pack(side=TOP, fill=X, padx=(10,0))
            # print(f"job => {job}")
            nameOfColumn = customtkinter.CTkTextbox(master=filtersContainerFixFrame, corner_radius=0, width=300, fg_color="white", text_color="black", border_spacing=0, border_color=self.color,
                                                        border_width=0, height=1, bg_color="blue", font=('Helvetica', 12, 'bold')) # , style="success.TEntry"
                    
            nameOfColumn.insert(END, job[0])
            nameOfColumn.configure(state=DISABLED)
            nameOfColumn.pack(side=LEFT, fill=X)
            
            filterOnColumn = customtkinter.CTkTextbox(master=filtersContainerFixFrame, corner_radius=0, width=2000, fg_color="white", text_color="black", border_spacing=0, border_color=self.color,
                                                            border_width=0, height=1, bg_color="blue", font=('Helvetica', 12, 'bold')) # , style="success.TEntry"
            if (job[1] == 1):
                filterOnColumn.insert(END, "True")
            elif (job[1] == 0):
                filterOnColumn.insert(END, "False")
            filterOnColumn.configure(state=DISABLED)
            filterOnColumn.pack(side=LEFT, fill=X)

        # self.skippedLogUploadUnique = self.skippedLogUpload

        # # for index, curList in enumerate(self.skippedLogUploadUnique):
        # #     if len(curList) > 6:
        # #         self.skippedLogUploadUnique = self.skippedLogUploadUnique.pop(index)

        # # print("self.skippedLogUploadUnique => ", self.skippedLogUploadUnique)

        # # for tuple in columnsWithFilters:
        # for key in self.resultSummary.keys():
        #     if key == "Upload done" or key == "Pending upload" or key == "Name didn't match":
        #         continue
        #     else:
        #         # # print("key in else => ", key)
        #         filtersContainerFrame = customtkinter.CTkFrame(self.currentFiltersFrame, border_width=0.1, corner_radius=1, fg_color=self.color, height=250)
        #         filtersContainerFrame.pack(side=TOP, fill=BOTH, padx=(10,0))
        #         nameOfColumn = customtkinter.CTkTextbox(master=filtersContainerFrame, corner_radius=0, width=150, fg_color="white", text_color="black", border_spacing=0, border_color=self.color,
        #                                                 border_width=0, height=1, bg_color="white", font=('Helvetica', 12)) # , style="success.TEntry"
                    
        #         nameOfColumn.insert(END, key)
        #         nameOfColumn.configure(state=DISABLED)
        #         nameOfColumn.pack(side=LEFT, fill=X)
                
        #         filterOnColumn = customtkinter.CTkTextbox(master=filtersContainerFrame, corner_radius=0, width=2000, fg_color="white", text_color="black", border_spacing=0, border_color=self.color,
        #                                                         border_width=0, height=1, bg_color="white", font=('Helvetica', 12)) # , style="success.TEntry"

        #         print("(self.resultSummary[key] - len(self.skippedLogUploadUnique)) => ", (self.resultSummary[key] - len(self.skippedLogUploadUnique)))
        #         if (self.resultSummary[key] - len(self.skippedLogUploadUnique)) > 0:
        #             filterOnColumn.insert(END, (self.resultSummary[key] - len(self.skippedLogUploadUnique)))
        #         elif key == "Skipped":
        #             filterOnColumn.insert(END, len(self.skippedLogUploadUnique))
        #         else:
        #             filterOnColumn.insert(END, self.resultSummary[key])
        #         filterOnColumn.configure(state=DISABLED)
        #         filterOnColumn.pack(side=LEFT, fill=X)
    

    def configureForPassword(self, event):
        # print("configureForPassword called...")
        if (self.passwordEntry.get() == "Enter Password..."):
            self.passwordEntry.delete(0, END)
        self.passwordEntry.configure(show="*")
    
    def configureAfterPassword(self, event):
        # print("configureAfterPassword called...")
        if (self.passwordEntry.get() == ""):
            self.passwordEntry.configure(show="")
            self.passwordEntry.insert(0, "Enter Password...")

    def storeTestPlanPath(self, event):
        self.pathToStoreEntry.delete(0, END)
        path = askdirectory()
        path = r'{}'.format(path)
        self.storeTestPlanLoc.set(path)
        self.pathToStoreEntry.insert(0, path)
    
    
    # DESC: Move Row Up
    def moveRecordUpFunction(self):
        self.selectedItem = self.treeView.selection() # get list of selected items e.g.;Item selected ->  ('I002', 'I003', 'I004', 'I005', 'I006', 'I007')

        for row in self.selectedItem:
            self.treeView.move(row, self.treeView.parent(row), self.treeView.index(row)-1)

    # DESC: Move Rown Down
    def moveRecordDownFunction(self):

        rows = self.treeView.selection()
        for row in reversed(rows):
            self.treeView.move(row, self.treeView.parent(row), self.treeView.index(row)+1)
       
 
    def resource_path(self, relative_path):
            """ Get absolute path to resource, works for dev and for PyInstaller """
            try:
                # PyInstaller creates a temp folder and stores path in _MEIPASS
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, relative_path)

    def restartLogUpload(self):
        self.uploadDoneFor.clear()
        self.couldNotUploadLogs.clear()
        self.allRecords.clear()
        self.skippedLogUpload.clear()
        self.testNameMismatchPair.clear()
        self.processedRecords.clear()
        self.startLogUploadButton.configure(command=self.startLogUpload)

    def itemSelected(self, event):
        self.treeView.bind('<Control-Shift-ButtonRelease-1>', self.controlShiftPressed)
        self.replaceStringInSelectedRecordsWindow = None
        self.replaceStringInSelectedColumnWindow = None
        self.clearContentOfSelectedColumnWindow = None
        self.selectedItem = self.treeView.selection() # get list of selected items e.g.;Item selected ->  ('I002', 'I003', 'I004', 'I005', 'I006', 'I007')

        
        self.listOfAllSelectedRecords = []
        for item in self.selectedItem:
            self.listOfAllSelectedRecords.append(self.treeView.item(item)['values'])
                
        if len(self.listOfAllSelectedRecords) >= 1:
            self.treeView.bind('<Control-r>', self.replaceStringInSelectedRecords)
        
    
    def controlShiftPressed(self, event):
        # self.selectedRow += 1
        ItemsToBeSelected = []
        # pass # pass # pass ## print("controlShiftPressed")
        pass # pass # pass ## print(self.treeView.selection())
        toBeSelectedRecords = self.treeView.get_children()
        # pass # pass # pass ## print("toBeSelectedRecords => ", toBeSelectedRecords)
        for item in self.treeView.selection():
            ItemsToBeSelected.append(item)
        # self.treeView.selection_add(item)
        
        firstIndex = ItemsToBeSelected[0]
        lastIndex = ItemsToBeSelected[-1]

        for index, item in enumerate(toBeSelectedRecords):
            if item == firstIndex:
                firstIndex = index
            elif item == lastIndex:
                lastIndex = index

        # pass # pass # pass ## print("firstIndex => ", firstIndex)
        # pass # pass # pass ## print("lastIndex => ", lastIndex)

        for index, item in enumerate(toBeSelectedRecords):
            if index >= firstIndex and index <= lastIndex:
                 self.treeView.selection_add(item)
    
    def pasteInTestName(self):

        self.replaceStringInSelectedRecordsWindow = None

        # df = pd.DataFrame(index=False)
        # dataFrame = pd.DataFrame()
        df = pd.DataFrame(columns=['TestName'])
        df = pd.read_clipboard(header=None,engine='python',
                            index_col=None,
                            encoding='utf-8')
        blankIndex=[''] * len(df)
        df.index=blankIndex
        df = df.fillna('')
        # pass ## print(df)
        dataFrame = pd.DataFrame(columns=['col1', 'TestName', 'Log path', 'Result', 'col2'])
        dataFrame = dataFrame.fillna('')
        # dataFrame.columns = ['SrNo', 'TestName']
        # print(df)
        dataFrame['TestName'] = df
        dataFrame[['col1','Log path','Result', 'col2']] = dataFrame[['col1', 'Log path','Result', 'col2']].fillna('')

        # print("dataFrame after removing null\n", dataFrame)
        pass ## print(f"{len(dataFrame)}")

        # pass ## print(dataFrame)
        # pass ## print(f"{len(dataFrame)=}")

        if (len(dataFrame)) > 1:
            self.treeView.bind('<Control-r>', self.replaceStringInSelectedRecords)

        indexColumn = []
        pass # pass ## print("dataFrame length -> ", len(dataFrame))
        for i in range(1, len(dataFrame) + 1):
            indexColumn.append(i)
        
        dataFrame.insert(1, "SrNo.", indexColumn, True)
        # dataFrame.columns = ['SrNo', 'TestName']
        # dataFrame.insert(1, "", "", True)

        pass ## print(dataFrame)
        pass ## print(f"{len(dataFrame)}")

        df_rows = dataFrame.to_numpy().tolist()

        self.treeView["column"] = self.columns
        self.treeView["show"] = "headings"

        for column in self.treeView["columns"]:
            self.treeView.heading(column, text=column)
        
        for column in self.treeView["columns"]:
            self.treeView.column(column, minwidth=300, width=300, anchor=CENTER, stretch=False)
        
        self.treeView.tag_configure("oddrow", background='white',foreground='black')
        self.treeView.tag_configure("evenrow", background='#dbeafe',foreground='black')

        for index, row in enumerate(df_rows):
            if index % 2 == 0:
                self.treeView.insert("", "end", values=row, tags=('evenrow',))
            if index % 2 != 0:
                self.treeView.insert("", "end", values=row, tags=('oddrow',))

        self.treeView.heading("#0", text="\n")
        self.treeView.column("#1", width=5, minwidth=5, anchor=CENTER, stretch=False)
        self.treeView.column("#2", width=100, minwidth=100, anchor=CENTER, stretch=False)
        self.treeView.column("#3", width=300, minwidth=300, anchor=CENTER, stretch=False)
        self.treeView.column("#4", width=1480, minwidth=1480, anchor=CENTER, stretch=False)
        self.treeView.column("#5", width=5, minwidth=5, anchor=CENTER, stretch=False)

        treeContent = self.treeView.get_children()

        if len(treeContent) > 0:
            self.importTemplateButton.configure(text="⤓ Save to Template")
            self.importTemplateButton.configure(command=lambda : self.saveDataToTemplateForUpload())

    def saveDataToTemplateForUpload(self):
        self.excel_filename = filedialog.askopenfilename(initialdir = "/",
                                        title = "Select a template file",
                                        filetypes = (("Excel files","*.xlsx*"), ("all files","*.*")))
        
        try:
            wb = openpyxl.load_workbook(self.excel_filename) 
        except:
            return
        
        treeViewContent = self.treeView.get_children()

        allRowsData = []


        for record in treeViewContent:
            allRowsData.append(self.treeView.item(record)['values'])
        
        print("allRowsData => ", treeViewContent)

        processedDataToSave = []

        for data in allRowsData:
            processedDataToSave.append(data[2:5])

        print("processedDataToSave=>", processedDataToSave)
        # print("processedDataToSave[0]=>", processedDataToSave[0])
        # print(len(processedDataToSave)-1)
        # print(len(processedDataToSave[0]))

        activeSheet = wb['Main_Upload']
        sh = activeSheet

        if self.targetNameEntry.get() != "Enter serial name for the host as per axiom database (e.g., N10RH6XXX)":
            targetNameCell = sh.cell(row=1, column=2)
            targetNameCell.value = self.targetNameEntry.get()
        
        for row in range(3, sh.max_row+1):
            for col in range(1, sh.max_column+1):
                cell_obj = sh.cell(row=row, column=col)
                cell_obj.value = ""

        index = 0
        for row in range(3, sh.max_row+1):
            for col in range(1, sh.max_column+1):
                if index < (len(processedDataToSave)) and (col-1) < (len(processedDataToSave[0])):
                    # print("index => ", index, " col => ", col-1)
                    cell_obj = sh.cell(row=row, column=col)
                    cell_obj.value = processedDataToSave[index][col-1]
            index += 1
        
        try:
            with open(self.excel_filename, "r+") as f:
                pass # print("opened file successfully to write")
            wb.save(self.excel_filename)
            messagebox.showinfo("Information", "Data saved successfully")
            # self.reloadTreeView()
        except IOError:
            messagebox.showinfo("Information", "Can't write to excel, check if file is open in another program...") 

    def deleteTestRecords(self, event):
        self.treeView.delete(*self.treeView.selection())
        treeViewLength = self.treeView.get_children()

        # print(len(treeViewLength))

        if len(treeViewLength) == 0 or len(treeViewLength) == len(self.uploadDoneFor):
            self.uploadDoneFor.clear()
            self.couldNotUploadLogs.clear()
            self.allRecords.clear()
            self.skippedLogUpload.clear()
            self.testNameMismatchPair.clear()
            self.processedRecords.clear()
            self.importTemplateButton.configure(text="⮯ Import from Template")
            self.importTemplateButton.configure(command=lambda : self.importTestFromExcel())
            # self.startLogUploadButton.configure(state="normal")
            self.index = 0


    def manipulateColumnOfSheet(self, event):
        region = self.treeView.identify("region", event.x, event.y)
        if region == "cell":
            pass # pass # pass ## print("allRecordsSelected false => ", self.allRecordsSelected)
            self.menuForColumn = ttk.Menu(self.treeView, tearoff=False)
             # command=lambda currentEntryInFocus = currentEntryInFocus: self.copy_select(currentEntryInFocus
            # self.menuForColumn.add_command(label="📝 Generate Script...", font = ("", 12), command= lambda : self.callToGenerateScript(event))
            # self.menuForColumn.add_separator()
            self.menuForColumn.add_command(label="✀ Delete...", font = ("", 12), command= lambda : self.deleteTestRecords(event))
            # self.menuForColumn.add_command(label="⇧ Paste Above...", font = ("", 12), command= lambda : self.pasteAboveSelectedRecord(event))
            # self.menuForColumn.add_command(label="⇩ Paste Below...", font = ("", 12), command= lambda : self.pasteBelowSelectedRecord(event))
            # self.menuForColumn.add_command(label="⌦ Clear column data...", font = ("", 12), command= lambda : self.clearColumnContent(event))
            # # self.menuForColumn.add_command(label="⥯ Replace...", font = ("", 12), command= lambda : self.replaceStringInSelectedRecords(event))
            # self.menuForColumn.add_command(label="⥯ Replace In Column...", font = ("", 12), command= lambda : self.replaceStringInSelectedColumn(event))
            self.menuForColumn.post(event.x_root, event.y_root)
    
        treeViewLength = self.treeView.get_children()

        if len(treeViewLength) == 0 and region == "nothing":
            self.menuForColumn = ttk.Menu(self.treeView, tearoff=False)
            self.menuForColumn.add_command(label="⎘  Paste Job IDs...", font = ("", 15), command= lambda : self.pasteInTestName())
            self.menuForColumn.post(event.x_root, event.y_root)

    """ Update data without going to update window"""
    def handleDoubleClick(self, event):
        row = self.treeView.selection()[0]
        jobIdName = self.treeView.item(row)['values'][2]

        column = self.treeView.identify_column(event.x)
        print(f"jobIdName => {jobIdName}")

        if column == "#4":
            print(f"self.summaryForJob => {self.summaryForJob}")
            self.viewJobSummary(self.summaryForJob[jobIdName], jobIdName)


    """ Select all rows functionality"""
    def selectAll(self, event):
        pass # pass ## print("Select all")
        self.replaceStringInSelectedRecordsWindow = None
        count = 0
        self.listOfAllSelectedRecords = []
        for item in self.treeView.get_children():
            self.selectAllRecords(item)
            self.listOfAllSelectedRecords.append(self.treeView.item(item)['values'])
            count += 1
        self.totalSelectedRecords = count
        pass # pass ## print("self.listOfAllSelectedRecords => ", self.listOfAllSelectedRecords)
        # if len(self.listOfAllSelectedRecords) > 1:
        #     self.treeView.bind('<Control-r>', self.replaceStringInSelectedRecords)
    
    def selectAllRecords(self, item):
        self.allRecordsSelected = True
        self.treeView.selection_add(item)
        
    def disableSelectedRecords(self, event):
        self.allRecordsSelected = False

    def replaceStringInSelectedRecords(self, event):
        # pass ## print(f"{self.replaceStringInSelectedRecordsWindow=}")
        if self.replaceStringInSelectedRecordsWindow is None:
            pass # pass ## print("replaceStringInSelectedRecords called")    

            self.replaceStringInSelectedRecordsWindow = tk.Toplevel()
            self.replaceStringInSelectedRecordsWindow.wm_overrideredirect(1)
            # self.replaceStringInSelectedRecordsWindow.wm_attributes("-topmost", True) 
            # self.renameSelectedColumnWindow.protocol("WM_DELETE_WINDOW", self.toplevelWindowClosed)
            windowToBeDestroyed = self.replaceStringInSelectedRecordsWindow
            # self.renameSelectedColumnWindow.title(columnName)
            # self.serachContentWindow.overrideredirect(True)
            self.replaceStringInSelectedRecordsWindow.resizable(0,0) 
            self.replaceStringInSelectedRecordsWindow.configure(bg="white", border=1, relief=SOLID)
            self.replaceStringInSelectedRecordsWindow.geometry("400x300+700+300")

            self.style.configure('custom.TLabel', font=('Roberto', 12, 'bold'))

            self.ContainerFrame = customtkinter.CTkFrame(self.replaceStringInSelectedRecordsWindow, border_width=0, corner_radius=0, fg_color="#3C005E")
            self.ContainerFrame.place(relx=0, rely=0, relwidth=1, relheight=0.35)

            font = customtkinter.CTkFont('Roberto', 14, 'bold')

            EntryLabel = customtkinter.CTkLabel(master=self.ContainerFrame, text="Replace", 
                                                        bg_color='transparent', text_color='white',
                                                        font=font)
            # SearchBoxEntryLabel.pack(side="top", expand=1, fill=X)
            EntryLabel.place(relx=0, rely=0.3, relwidth=1)

            self.style.configure('custom.TButton', background='#3C005E', foreground='white', font=('Helvetica', 12))

            EntryObj = PlaceholderEntry(self.replaceStringInSelectedRecordsWindow, placeholder="Find what...")
            EntryObj = EntryObj.e
            # EntryObj.insert(0, "Find what...")
            EntryObj.configure(style='info.TEntry')

            EntryObj.place(relx=0.08, rely=0.45, relwidth=0.85)

            EntryObj2 = PlaceholderEntry(self.replaceStringInSelectedRecordsWindow, placeholder="Replace with...")
            EntryObj2 = EntryObj2.e
            # EntryObj2.insert(0, "Replace with...")
            EntryObj2.configure(style='info.TEntry')

            EntryObj2.place(relx=0.08, rely=0.63, relwidth=0.85)

            replaceButton = ttk.Button(master=self.replaceStringInSelectedRecordsWindow, 
                                        width=15, text="Replace", 
                                        style='custom.TButton', 
                                        command=lambda EntryObj=EntryObj, EntryObj2=EntryObj2 : self.replaceRecordsInTreeView(EntryObj, EntryObj2))
            replaceButton.pack(side=LEFT, padx=(30, 0), pady=(220, 0))

            closeButton = ttk.Button(master=self.replaceStringInSelectedRecordsWindow, 
                                        width=15, text="Close", 
                                        style='custom.TButton', 
                                        command= lambda windowToBeDestroyed=windowToBeDestroyed : self.closeTopLevelWindow(windowToBeDestroyed))
            closeButton.pack(side=LEFT, padx=(22, 0), pady=(220, 0))
        
        else:
            # try:
            self.replaceStringInSelectedRecordsWindow.deiconify()
            # except:
            #     pass

    def replaceRecordsInTreeView(self, findEntry, ReplaceEntry):

        selectedItems = self.treeView.selection()

        # treeContentForSaveEdit = self.treeView.get_children()

        # pass # pass ## print("findEntry => ", findEntry.get().strip())
        if findEntry.get().isdigit():
            findString = findEntry.get()
        else:
            findString = findEntry.get().strip()
        # pass # pass ## print("ReplaceEntry => ", ReplaceEntry.get().strip())
        if ReplaceEntry.get().isdigit():
            replaceString = ReplaceEntry.get()
        else:
            replaceString = ReplaceEntry.get().strip()
        pass ## print("self.listOfAllSelectedRecords => ", self.listOfAllSelectedRecords)

        self.srNoOfRecordsToReplaceStringIn = []

        self.listOfAllSelectedRecordsWithReplacedString = []

        for record in self.listOfAllSelectedRecords:
            self.srNoOfRecordsToReplaceStringIn.append((record[0] + 1))
        
        self.foundMatch = False

        numberOfReplacementsMade = 0

        for record in self.listOfAllSelectedRecords:
            RecordsWithReplacedString = []
            for value in record:
                # pass # pass ## print("value =>", value)
                if type(value) is str and findString in value:
                    pass # pass ## print("String found..")
                    self.foundMatch = True
                    numberOfTimesStringFoundInRecord = value.count(findString)
                    numberOfReplacementsMade += numberOfTimesStringFoundInRecord
                    pass # pass ## print("Old data => ", value)
                    value = value.replace(findString, replaceString)
                    pass # pass ## print("Updated data => ", value)
                    RecordsWithReplacedString.append(value)
                else:
                    RecordsWithReplacedString.append(value)
            
            pass # pass ## print("RecordsWithReplacedString => ", RecordsWithReplacedString)
            if len(RecordsWithReplacedString) != 0:
                self.listOfAllSelectedRecordsWithReplacedString.append(RecordsWithReplacedString)

        if self.foundMatch == False:
            messagebox.showerror("Information", "No matching record found with current string, please check string again...")
            return

        for index, row in enumerate(selectedItems):
            self.treeView.item(row, text='', values=self.listOfAllSelectedRecordsWithReplacedString[index])
        

    def closeTopLevelWindow(self, windowToBeDestroyed):
        self.replaceStringInSelectedRecordsWindow = None
        windowToBeDestroyed.destroy()
    
    def closeTopLevelWindowColumn(self, windowToBeDestroyed):
        self.replaceStringInSelectedColumnWindow = None
        # for i in self.treeView.selection():
        #     self.treeView.selection_remove(i)
        windowToBeDestroyed.destroy()
    

    def replaceStringInSelectedColumn(self, event):
        if self.replaceStringInSelectedColumnWindow is None:
            pass # pass # pass ## print("replaceStringInSelectedColumnWindow called")    
            # currentActiveTab=str(self.tabview.tab(self.tabview.select(), "text"))
            # currentActiveTab = currentActiveTab.strip()

            self.replaceStringInSelectedColumnWindow = tk.Toplevel()
            self.replaceStringInSelectedColumnWindow.wm_overrideredirect(1) 
            # self.renameSelectedColumnWindow.protocol("WM_DELETE_WINDOW", self.toplevelWindowClosed)
            windowToBeDestroyed = self.replaceStringInSelectedColumnWindow
            # self.renameSelectedColumnWindow.title(columnName)
            # self.serachContentWindow.overrideredirect(True)
            self.replaceStringInSelectedColumnWindow.resizable(0,0) 
            self.replaceStringInSelectedColumnWindow.configure(bg="white", border=1, relief=SOLID)
            self.replaceStringInSelectedColumnWindow.geometry("400x350+700+300")

            self.style.configure('custom.TLabel', font=('Roberto', 12, 'bold'))

            self.ContainerFrame = customtkinter.CTkFrame(self.replaceStringInSelectedColumnWindow, border_width=0, corner_radius=0, fg_color="#3C005E")
            self.ContainerFrame.place(relx=0, rely=0, relwidth=1, relheight=0.35)

            font = customtkinter.CTkFont('Roberto', 14, 'bold')

            EntryLabel = customtkinter.CTkLabel(master=self.ContainerFrame, text="Replace in column", 
                                                        bg_color='transparent', text_color='white',
                                                        font=font)
            # SearchBoxEntryLabel.pack(side="top", expand=1, fill=X)
            EntryLabel.place(relx=0, rely=0.3, relwidth=1)

            # EntryObj3 = PlaceholderEntry(self.replaceStringInSelectedColumnWindow, placeholder="Column Name")
            # EntryObj3 = EntryObj3.e
            # # EntryObj.insert(0, "Find what...")
            # EntryObj3.configure(style='info.TEntry')

            # self.style.configure("TCombobox", fieldbackground= "#3C005E", background= "white")

            # self.style.configure('custom.TCombobox', background='self.color', foreground='white', font=('Helvetica', 12))


            EntryObj3 = ttk.Combobox(self.replaceStringInSelectedColumnWindow,
                        state="readonly", font=('Helvetica', 12),
                        values=["Select column name...", "Test Name", "Log path", "Result"], style='info.TCombobox')

            EntryObj3.current(0)
            EntryObj3.place(relx=0.08, rely=0.4, relwidth=0.85)

            EntryObj = PlaceholderEntry(self.replaceStringInSelectedColumnWindow, placeholder="Find what...")
            EntryObj = EntryObj.e
            # EntryObj.insert(0, "Find what...")
            EntryObj.configure(style='info.TEntry')

            EntryObj.place(relx=0.08, rely=0.52, relwidth=0.85)

            EntryObj2 = PlaceholderEntry(self.replaceStringInSelectedColumnWindow, placeholder="Replace with...")
            EntryObj2 = EntryObj2.e
            # EntryObj2.insert(0, "Replace with...")
            EntryObj2.configure(style='info.TEntry')

            EntryObj2.place(relx=0.08, rely=0.64, relwidth=0.85)

            self.style.configure('custom.TButton', background='#3C005E', foreground='white', font=('Helvetica', 12))

            replaceButton = ttk.Button(master=self.replaceStringInSelectedColumnWindow, 
                                        width=14, text="Replace", 
                                        style='custom.TButton', 
                                        command=lambda EntryObj=EntryObj, EntryObj2=EntryObj2 : self.replaceRecordsInColumn(EntryObj, EntryObj2, EntryObj3))
            replaceButton.pack(side=LEFT, padx=(35, 0), pady=(250, 0))

            closeButton = ttk.Button(master=self.replaceStringInSelectedColumnWindow, 
                                        width=14, text="Close", 
                                        style='custom.TButton', 
                                        command= lambda windowToBeDestroyed=windowToBeDestroyed : self.closeTopLevelWindowColumn(windowToBeDestroyed))
            closeButton.pack(side=LEFT, padx=(30, 0), pady=(250, 0))
        
        else:
            try:
                self.replaceStringInSelectedColumnWindow.deiconify()
            except:
                pass
    
    
    def replaceRecordsInColumn(self, findEntry, ReplaceEntry, columnNameEntry):
        
        selectedItems = self.treeView.selection()
        
        tmpList = []
        if findEntry.get().isdigit():
            findString = findEntry.get()
        else:
            findString = findEntry.get().strip()
        # pass ## print("ReplaceEntry => ", ReplaceEntry.get().strip())
        if ReplaceEntry.get().isdigit():
            replaceString = ReplaceEntry.get()
        else:
            replaceString = ReplaceEntry.get().strip()
        
        if columnNameEntry.get().isdigit():
            columnName = columnNameEntry.get()
        else:
            columnName = columnNameEntry.get().strip()
        # columnName = columnNameEntry.get().strip()


        pass ## print(f"{columnName}")

        if columnName == 'Select column name...':
            messagebox.showerror('Error', "Please select a valid column name to continue")
            return
        
        tmpList = []
        columnNameMultipleString = columnName.split(",")
        # findMultipleString = x.strip() 
        for x in columnNameMultipleString:
            if x is not None:
                tmpList.append(x.strip())

        columnNameMultipleString = tmpList
        lenOfColumnNameMultipleString = len(columnNameMultipleString)
        pass ## print("columnNameMultipleString => ", columnNameMultipleString)

        tmpList = []
        if "Find what..." not in findString:
            findMultipleString = findString.split(",")
            pass ## print(f"{findMultipleString}")
            # findMultipleString = x.strip() 
            for x in findMultipleString:
                if x is not None and x != '':
                    tmpList.append(x.strip())
            findMultipleString = tmpList
            lenOfFindMultipleString = len(findMultipleString)
            pass ## print("findMultipleString => ", findMultipleString)
        
        # if lenOfFindMultipleString != lenOfColumnNameMultipleString:

        # else:
        #     pass ## print("findString => ", findString)

        tmpList = []
        replaceMultipleString = replaceString.split(",")
        # findMultipleString = x.strip() 
        for x in replaceMultipleString:
            if x is not None:
                tmpList.append(x.strip())

        replaceMultipleString = tmpList
        lenOfReplaceMultipleString = len(replaceMultipleString)
        pass ## print("replaceMultipleString => ", replaceMultipleString)

        
        if columnName == '':
            messagebox.showerror("Error", "Please enter column name to continue")
            return

        if replaceString == '':
            messagebox.showerror("Error", "Please enter replace string to continue")
            return
        
        pass ## print("self.listOfAllSelectedRecords => ", self.listOfAllSelectedRecords)

        
        itr = 0
        if "Find what..." in findString:
            findMultipleString = []
            while itr < lenOfColumnNameMultipleString:
                findMultipleString.append("")
                itr += 1
            lenOfFindMultipleString = len(findMultipleString)
            pass ## print("findMultipleString for empty entry and multi column mode => ", findMultipleString)

        if lenOfReplaceMultipleString != lenOfColumnNameMultipleString or lenOfFindMultipleString != lenOfColumnNameMultipleString:
            messagebox.showerror("Information", "Please make sure to enter data for all columns")
            return

        findMultipleString = findMultipleString[0]
        replaceMultipleString = replaceMultipleString[0]
        columnNameMultipleString = columnNameMultipleString[0]
        srNoOfRecordsToReplaceStringIn = []

        totalFoundStringFor = 0

        numberOfReplacementsMade = 0


        if columnName.lower().strip() in self.columnslowerCase:
            colIndex = self.columnslowerCase.index(columnName.lower())
        else:
            messagebox.showerror("Error", "Please enter a proper column name")
            return

        for index, record in enumerate(self.listOfAllSelectedRecords):
            pass ## print("Current record => ", record)
            self.foundMatch = False
            if type(record[colIndex]) == str and findString in record[colIndex]:
                # pass # pass ## print("String found..")
                self.foundMatch = True
                # numberOfReplacementsMade += 1
                numberOfTimesStringFoundInRecord = record[colIndex].count(findString)
                numberOfReplacementsMade += numberOfTimesStringFoundInRecord
                value = record[colIndex].replace(findString, replaceString)
                self.listOfAllSelectedRecords[index][colIndex] = value
                # RecordsWithReplacedString.append(record[colIndex])
            if findString == 'Find what...' or findString == '' and record[colIndex] == '':
                self.foundMatch = True
                numberOfReplacementsMade += 1
                value = replaceString
                self.listOfAllSelectedRecords[index][colIndex] = value
            if self.foundMatch == True:
                # numberOfReplacementsMade += 1
                totalFoundStringFor += 1

        for index, row in enumerate(selectedItems):
            self.treeView.item(row, text='', values=self.listOfAllSelectedRecords[index])

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        """ Set window full screen intially """
        self._state_before_windows_set_titlebar_color = 'zoomed'


        self.style = ttk.Style()

        self.style.theme_use('pulse')

        self.wm_protocol("WM_DELETE_WINDOW", self.closeAllWindows)

        self.frames = [homeWindow(self), getXunitFiles(self)]
        # color = "#0d6efd"
        color = "#0d6efd"

        # blueColor = #1d4ed8

        self.style.configure("custom.TButton", font=('Roboto', 11), 
                             background=color, focuscolor="green3", 
                             bordercolor=color, relief=FLAT)

        # ModifyExcelWindow(self)

        # createExcelWindow(self)

        self.grid_columnconfigure(1, weight=1)
        # self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=8)

        self.title("X-Unit Download")
        self.geometry("1520x780+0+0")

        def resource_path(relative_path):
            """ Get absolute path to resource, works for dev and for PyInstaller """
            try:
                # PyInstaller creates a temp folder and stores path in _MEIPASS
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, relative_path)

            # self.selectAxiomParamsWindow.iconbitmap(resource_path('icon.ico'))

        # img = ImageTk.PhotoImage(Image.open(resource_path("logoForAxiomLogUpload.png")))


        #menu frame
        self.menuFrame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="#0d6efd")
        self.menuFrame.place(relx=0, rely=0, relwidth=1, relheight=0.075)

        # Create a Label Widget to display logo
        # label = customtkinter.CTkLabel(self.menuFrame, image = img, text="")
        # label.pack(side=LEFT, padx=(10, 0), pady=(8, 5))

        self.menuBtnContainerOne = customtkinter.CTkFrame(self.menuFrame, corner_radius=0, fg_color="#0d6efd")
        self.menuBtnContainerOne.pack(side=LEFT, padx=(10, 0), pady=(8, 5))

        self.menuBtnContainerTwo = customtkinter.CTkFrame(self.menuFrame, corner_radius=0, fg_color="#0d6efd")
        self.menuBtnContainerTwo.pack(side=LEFT, padx=(10, 0), pady=(8, 5))
        
        self.menuBtnContainerThree = customtkinter.CTkFrame(self.menuFrame, corner_radius=0, fg_color="#0d6efd")
        self.menuBtnContainerThree.pack(side=LEFT, padx=(10, 0), pady=(8, 5))

        self.menuBtnContainerFour = customtkinter.CTkFrame(self.menuFrame, corner_radius=0, fg_color="#0d6efd")
        self.menuBtnContainerFour.pack(side=LEFT, padx=(10, 0), pady=(8, 5))

        # homeWindowClicked = "homeWindowClicked"
        # self.homeWindowBtn = customtkinter.CTkButton(master=self.menuBtnContainerTwo, width=50, height=40, text="Home", 
        #                                             fg_color=color, hover_color="#663399", image=customtkinter.CTkImage(light_image=Image.open("./Icons/homeIcon.png"), size=(20, 15)),
        #                                             corner_radius=10, border_width=0, anchor="w", 
        #                                             font=customtkinter.CTkFont('Helvetica', 15, 'bold'),
        #                                             command=lambda homeWindowClicked=homeWindowClicked : self.raiseFrame(homeWindowClicked))
        # self.homeWindowBtn.pack(side=TOP, pady=(1, 0))

        # generateScriptClicked = "generateScriptClicked"
        # self.generateScriptBtn = customtkinter.CTkButton(master=self.menuBtnContainerFour, width=50, height=40, 
        #                                                 text="GetXunitFiles", fg_color=color, hover_color="#663399",
        #                                                 corner_radius=10, border_width=0, anchor="w",
        #                                                 image=customtkinter.CTkImage(light_image=Image.open("./Icons/GenerateIcon.png"), size=(20, 15)),
        #                                                 font=customtkinter.CTkFont('Helvetica', 15, 'bold'),
        #                                             command=lambda generateScriptClicked=generateScriptClicked : self.raiseFrame(generateScriptClicked))
        # self.generateScriptBtn.pack(side=TOP, pady=(1, 0))

        getXunitFilesClicked = "getXunitFilesClicked"
        self.getXunitFilesBtn = customtkinter.CTkButton(master=self.menuBtnContainerOne, width=50, height=40, 
                                                        text="DownloadXunitFiles", fg_color=color, hover_color="#663399",
                                                        corner_radius=10, border_width=0, anchor="w", 
                                                        font=customtkinter.CTkFont('Helvetica', 15, 'bold'),
                                                        command=lambda getXunitFilesClicked=getXunitFilesClicked : self.raiseFrame(getXunitFilesClicked))
        self.getXunitFilesBtn.pack(side=TOP, pady=(1, 0))

        # ExecuteScriptAndUploadLogsClicked = "ExecuteScriptAndUploadLogsClicked"
        # self.ExecuteScriptAndUploadLogsBtn = customtkinter.CTkButton(master=self.menuBtnContainerTwo, width=50, height=40, 
        #                                                             text="ExecuteTestScripts", fg_color=color, hover_color="#663399",
        #                                                             image=customtkinter.CTkImage(light_image=Image.open("./Icons/ExcuteIcon.png"), size=(20, 15)),
        #                                                             corner_radius=10, border_width=0, anchor="w", font=customtkinter.CTkFont('Helvetica', 15, 'bold'),
        #                                                             command=lambda ExecuteScriptAndUploadLogsClicked=ExecuteScriptAndUploadLogsClicked : self.raiseFrame(ExecuteScriptAndUploadLogsClicked))
        # self.ExecuteScriptAndUploadLogsBtn.pack(side=TOP, pady=(1, 0))


    def closeAllWindows(self):
        pass ## print("closeAllWindows called")
        # global activeProcess
        if activeProcess != 'Null':
            poll = activeProcess.poll()
            if poll is None:
                print("subprocess is alive")
            else:
                print("Process is no longer running, starting next script")
        self.quit()

    def raiseFrame(self, frameClicked):
        color = "#008000"
        global frames
        # pass # pass ## print("Frames -> ", frames)

        # pass # pass ## print("Frame clicked -> ", frameClicked)

        # if frameClicked == "getXunitFilesClicked":
        #     # pass # pass ## print("inside if createClicked")
        #     # self.createExcelButtonIndicator.configure(background="white")
        #     # self.createExcelButton.configure(font=('Roboto', 15, 'bold'))
        #     # self.modifyButtonIndicator.configure(background=color)
        #     # self.getXunitFilesBtn.configure(font=customtkinter.CTkFont('Helvetica', 15, 'bold'))
        #     # self.after(100, None)
        #     # self.generateScriptBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
        #     # self.ExecuteScriptAndUploadLogsBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
        #     # self.homeWindowBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
            
        #     # self.generateTestPlanForSwe5Button.configure(font=('Roboto', 15))
        #     frames[0].forget()
        #     frames[3].forget()
        #     frames[2].forget()
        #     frames[1].tkraise()


        if frameClicked == "homeWindowClicked":
            # self.homeWindowBtn.configure(font=customtkinter.CTkFont('Helvetica', 15, 'bold'))
            # self.after(100, None)
            # self.generateScriptBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
            # self.getXunitFilesBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
            # self.ExecuteScriptAndUploadLogsBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
            
            # pass # pass ## print("inside if modifyClicked")
            frames[1].forget()
            # frames[2].forget()
            # frames[0].forget()
            frames[0].tkraise()

        # elif frameClicked == "ExecuteScriptAndUploadLogsClicked":
        #     # self.modifyButtonIndicator.configure(background="white")
        #     # self.ExecuteScriptAndUploadLogsBtn.configure(font=customtkinter.CTkFont('Helvetica', 15, 'bold'))
        #     # self.after(100, None)
        #     # self.homeWindowBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
        #     # self.generateScriptBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
        #     # self.getXunitFilesBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
            
        #     # pass # pass ## print("inside if modifyClicked")
        #     # frames[2].forget()
        #     frames[3].forget()
        #     frames[1].forget()
        #     frames[0].forget()
        #     frames[2].tkraise()

        elif frameClicked == "getXunitFilesClicked":
            # self.modifyButtonIndicator.configure(background="white")
            # self.generateScriptBtn.configure(font=customtkinter.CTkFont('Helvetica', 15, 'bold'))
            # self.after(100, None)
            # self.homeWindowBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
            # self.ExecuteScriptAndUploadLogsBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
            # self.getXunitFilesBtn.configure(font=customtkinter.CTkFont('Helvetica', 15))
            
            # self.generateTestPlanForSwe5Button.configure(font=('Roboto', 15))
            # self.createExcelButton.configure(font=('Roboto', 15))
            # pass # pass ## print("inside if modifyClicked")
            # frames[2].forget()
            # frames[3].forget()
            # frames[2].forget()
            frames[0].forget()
            frames[1].tkraise()
 


if __name__ == "__main__":
    app = App()
    # app.state('zoomed')
    # app.wm_state('zoomed')
    # icon = Image.open('./logoFeviconSmall.png')
    # icon = icon.save('icon.ico',format = 'ICO', sizes=[(32,32)])

    def resource_path(relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)
    
    app.iconbitmap(resource_path('icon.ico'))
    app.mainloop()